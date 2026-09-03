# -*- coding: utf-8 -*-
r"""
Prepara los CSV de carga a partir de los Excel reales del ASIS.

    python preparar-datos.py

Lee de ../../ los dos archivos descargados del ASIS y escribe datos/*.csv,
que es lo que consume 13-carga-asis.sql con \copy.

La carpeta datos/ esta en .gitignore: los CSV NO se versionan, porque
duplicarian los datos personales que ya estan en los Excel de origen.

NOTA SOBRE NOMBRES Y CORREOS
----------------------------
El informe actual del ASIS entrega solo tres columnas -Doc ID, ID y Prog Acad-.
No trae nombres ni correos. Este script los SINTETIZA a partir del ID para que
la base quede cargable, y lo deja marcado en la columna nombre_completo_asis.

Eso no es un atajo: es la evidencia del riesgo numero 1 del plan. Cuando Carlos
amplie el informe a seis columnas -nombre, correo institucional y correo
personal, acordado en 1:50:43-, basta con cambiar las tres funciones marcadas
con SINTETICO y volver a cargar.
"""
import csv
import os
import re
import sys
import unicodedata

import openpyxl

RAIZ = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
SALIDA = os.path.join(os.path.dirname(__file__), 'datos')
os.makedirs(SALIDA, exist_ok=True)

F_EVENTOS = os.path.join(RAIZ, 'USBME_LCONTROL_CATEDRAS_1076195337.xlsx')
F_PERSONAS = os.path.join(RAIZ, 'USBME_EMPLID_CATED_V1_1504983333.xlsx')


def escribir(nombre, cabecera, filas):
    ruta = os.path.join(SALIDA, nombre)
    with open(ruta, 'w', newline='', encoding='utf-8') as fh:
        w = csv.writer(fh)
        w.writerow(cabecera)
        w.writerows(filas)
    sys.stdout.write('  %-28s %7d filas\n' % (nombre, len(filas)))
    return len(filas)


def sin_tildes(t):
    return ''.join(c for c in unicodedata.normalize('NFD', t)
                   if unicodedata.category(c) != 'Mn')


# ===========================================================================
# 1 · Eventos y reuniones  ->  catedra + sesion
# ===========================================================================
sys.stdout.write('Leyendo eventos...\n')
wb = openpyxl.load_workbook(F_EVENTOS, data_only=True)
ws = wb.active

catedras = {}        # id_evento -> (descripcion, tipo)
sesiones = []        # (id_evento, numero_reunion)
descartadas = 0

for fila in ws.iter_rows(min_row=3, values_only=True):
    if not fila or not fila[0]:
        continue
    evento = str(fila[0]).strip()
    if not re.fullmatch(r'\d{1,9}', evento):
        descartadas += 1
        continue
    evento = evento.zfill(9)                      # H8 · ceros a la izquierda
    try:
        reunion = int(float(fila[1]))             # H12 · llega como 1.0
    except (TypeError, ValueError):
        descartadas += 1
        continue
    desc = (str(fila[2]).strip() if fila[2] else 'SIN DESCRIPCION')[:200]
    tipo = (str(fila[3]).strip().upper() if fila[3] else 'CAAB')[:4]

    if evento not in catedras:
        catedras[evento] = (desc, tipo)
    sesiones.append((evento, reunion))

wb.close()

# H4 · el par (evento, reunion) es unico; se comprueba en vez de suponerlo
vistos, sesiones_ok, dup = set(), [], 0
for ev, re_ in sesiones:
    if (ev, re_) in vistos:
        dup += 1
        continue
    vistos.add((ev, re_))
    sesiones_ok.append((ev, re_))

escribir('catedra.csv', ['id_evento_asis', 'nombre', 'tipo_evento'],
         [(k, v[0], v[1]) for k, v in sorted(public.items())])
escribir('sesion.csv', ['id_evento_asis', 'numero_reunion'], sorted(sesiones_ok))
sys.stdout.write('  descartadas=%d  pares duplicados=%d\n' % (descartadas, dup))

# ===========================================================================
# 2 · Personas  ->  asistente + documento + programa
# ===========================================================================
sys.stdout.write('Leyendo maestro de personas...\n')
wb = openpyxl.load_workbook(F_PERSONAS, data_only=True)
ws = wb.active

personas = {}        # id_asis -> set de documentos
matriculas = set()   # (id_asis, codigo_programa)
programas = set()
novedades = []       # filas rechazadas, con su motivo

for n, fila in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
    if not fila or not fila[1]:
        continue
    doc = str(fila[0]).strip() if fila[0] else ''
    emplid = str(fila[1]).strip()
    prog = str(fila[2]).strip().upper() if fila[2] else ''

    # H1 y H2 · alfanumerico de 5 a 15
    if not re.fullmatch(r'[0-9A-Za-z]{5,15}', emplid):
        novedades.append((n, '%s|%s|%s' % (doc, emplid, prog), 'ID fuera de dominio'))
        continue
    if doc and not re.fullmatch(r'[0-9A-Za-z-]{4,20}', doc):
        novedades.append((n, '%s|%s|%s' % (doc, emplid, prog), 'Documento fuera de dominio'))
        doc = ''
    if prog and not re.fullmatch(r'[A-Z][0-9A-Z]{3,4}', prog):
        novedades.append((n, '%s|%s|%s' % (doc, emplid, prog), 'Codigo de programa fuera de dominio'))
        prog = ''

    personas.setdefault(emplid, set())
    if doc:
        personas[emplid].add(doc)
    if prog:
        programas.add(prog)
        matriculas.add((emplid, prog))       # H11 · el set deduplica

wb.close()

# --- SINTETICO · hasta que el informe del ASIS traiga estas tres columnas ---
def nombres_sinteticos(emplid):
    return ('Asistente', 'ASIS %s' % emplid,
            'Nombre no entregado por el informe actual del ASIS')


def correo_sintetico(emplid):
    return 'asis.%s@usbmed.edu.co' % emplid.lower()


filas_asis = []
for emplid in sorted(personas):
    nom, ape, marca = nombres_sinteticos(emplid)
    filas_asis.append((emplid, nom, ape, marca, correo_sintetico(emplid)))
escribir('asistente.csv',
         ['id_asis', 'nombres', 'apellidos', 'nombre_completo_asis', 'correo_institucional'],
         filas_asis)

filas_doc = []
for emplid in sorted(personas):
    for i, doc in enumerate(sorted(personas[emplid])):
        # El mas largo se toma como vigente; los demas quedan como historicos
        filas_doc.append((emplid, 'CC', doc, 'true' if i == len(personas[emplid]) - 1 else 'false'))
escribir('documento.csv', ['id_asis', 'tipo', 'numero', 'vigente'], filas_doc)

escribir('programa.csv', ['codigo', 'nombre'],
         [(p, 'Programa %s' % p) for p in sorted(programas)])
escribir('matricula.csv', ['id_asis', 'codigo'], sorted(matriculas))
escribir('novedad.csv', ['numero_fila', 'contenido_crudo', 'motivo'], novedades)

# ===========================================================================
# 3 · Resumen, para contrastarlo con lo que dice el plan
# ===========================================================================
multi = sum(1 for v in personas.values() if len(v) > 1)
sys.stdout.write('\nResumen\n')
sys.stdout.write('  catedras            %6d\n' % len(catedras))
sys.stdout.write('  sesiones            %6d\n' % len(sesiones_ok))
sys.stdout.write('  personas            %6d\n' % len(personas))
sys.stdout.write('  con >1 documento    %6d   <- H10 del plan: deben ser 707\n' % multi)
sys.stdout.write('  programas           %6d   <- H12 del plan: deben ser 105\n' % len(programas))
sys.stdout.write('  matriculas          %6d\n' % len(matriculas))
sys.stdout.write('  novedades           %6d\n' % len(novedades))
